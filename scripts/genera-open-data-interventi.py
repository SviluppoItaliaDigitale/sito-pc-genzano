#!/usr/bin/env python3
"""Genera dataset OPEN DATA AGGREGATI dagli export del gestionale interventi.

Legge l'export .xlsx del gestionale (fogli: Interventi, Interventi membri,
Risposte Selettive) e produce SOLO statistiche aggregate e anonime in
static/open-data/. NON pubblica MAI dati personali o sensibili: nessun nome,
nessun richiedente, nessuna descrizione, nessun luogo/indirizzo specifico.

Il file .xlsx NON va committato (contiene nomi dei volontari): resta sul PC.
Questo script ne estrae solo i numeri.

Uso:  python3 scripts/genera-open-data-interventi.py [percorso.xlsx]

Senza argomento prende automaticamente l'export PIÙ RECENTE in ~/Scrivania
(file `gcpc-genzano-di-roma_global_reports_*.xlsx`): il nome cambia a ogni
esportazione, quindi basta scaricare il nuovo file e rilanciare lo script.

L'export del gestionale è CUMULATIVO ("global report"): contiene sempre tutto
dall'adozione del sistema. Perciò NON serve scartare a mano i dati già usati:
ogni nuovo export ricontiene i vecchi + i nuovi, e questo script ricalcola e
SOVRASCRIVE i totali aggregati. I dati restano PROVVISORI (dal nuovo gestionale,
non da inizio anno). Se un export NON fosse cumulativo, va usata una logica di
merge: in quel caso segnalarlo.
"""
import csv
import glob
import json
import re
import sys
import os
import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Serve openpyxl:  pip install --break-system-packages openpyxl")

GLOB = os.path.expanduser("~/Scrivania/gcpc-genzano-di-roma_global_reports_*.xlsx")
OUT = "static/open-data"


def trova_export():
    """Ritorna l'export più recente che corrisponde al pattern, o None."""
    cand = glob.glob(GLOB)
    return max(cand, key=os.path.getmtime) if cand else None

MESI = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5,
        "giugno": 6, "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10,
        "novembre": 11, "dicembre": 12}
MESI_NOME = {v: k for k, v in MESI.items()}

NOTA = ("Dati PROVVISORI: censiti solo dall'adozione del nuovo gestionale, "
        "non coprono l'intero anno (mancano i mesi precedenti). Aggregati e "
        "anonimi, senza dati personali.")


def parse_data_it(s):
    if not s:
        return None
    s = str(s).lower()
    m = re.search(r'(\d{1,2})\s+([a-zà]+)\s+(\d{4})', s)
    if not m:
        return None
    g, mese, a = int(m.group(1)), MESI.get(m.group(2)), int(m.group(3))
    if not mese:
        return None
    try:
        return datetime.date(a, mese, g)
    except ValueError:
        return None


def minuti(s):
    if s is None:
        return 0
    tot = 0
    for n, u in re.findall(r'(\d+)\s*(or[ae]|min\w*)', str(s).lower()):
        tot += int(n) * (60 if u.startswith("or") else 1)
    return tot


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def fnum(x):
    v = num(x)
    return v if v is not None else 0


def write_kv(nome, indicatori, periodo):
    """Scrive un dataset chiave/valore in CSV + JSON."""
    csv_path = os.path.join(OUT, nome + ".csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow(["indicatore", "valore"])
        for k, v in indicatori:
            w.writerow([k, v])
    json_path = os.path.join(OUT, nome + ".json")
    with open(json_path, "w", encoding="utf-8") as fp:
        json.dump({
            "dataset": nome,
            "fonte": "Gruppo Comunale Volontari di Protezione Civile di Genzano di Roma",
            "licenza": "CC BY 4.0",
            "periodo": periodo,
            "nota": NOTA,
            "generato_il": datetime.date.today().isoformat(),
            "dati": {k: v for k, v in indicatori},
        }, fp, ensure_ascii=False, indent=2)
    print(f"  scritto {csv_path} + .json")


def write_rows(nome, header, rows, periodo):
    csv_path = os.path.join(OUT, nome + ".csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow(header)
        w.writerows(rows)
    json_path = os.path.join(OUT, nome + ".json")
    with open(json_path, "w", encoding="utf-8") as fp:
        json.dump({
            "dataset": nome,
            "fonte": "Gruppo Comunale Volontari di Protezione Civile di Genzano di Roma",
            "licenza": "CC BY 4.0",
            "periodo": periodo,
            "nota": NOTA,
            "generato_il": datetime.date.today().isoformat(),
            "dati": [dict(zip(header, r)) for r in rows],
        }, fp, ensure_ascii=False, indent=2)
    print(f"  scritto {csv_path} + .json")


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else trova_export()
    if not src or not os.path.exists(src):
        sys.exit("Export non trovato. Passa il percorso del .xlsx come argomento, "
                 "oppure mettilo in ~/Scrivania col nome gcpc-...global_reports_*.xlsx")
    print(f"Export: {os.path.basename(src)}")
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    # ---------- INTERVENTI ----------
    ws = wb["Interventi"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr) if h}

    def col(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    data = [r for r in rows[1:] if col(r, "Numero") not in (None, "")]
    n_tot = len(data)
    min_tot = sum(minuti(col(r, "Durata")) for r in data)
    ore = round(min_tot / 60, 1)
    km = 0
    for r in data:
        a, b = num(col(r, "Km iniziali")), num(col(r, "Km finali"))
        if a is not None and b is not None and b >= a:
            km += b - a
    date = [parse_data_it(col(r, "Data Inizio")) for r in data]
    date = [d for d in date if d]
    periodo = {"dal": min(date).isoformat() if date else None,
               "al": max(date).isoformat() if date else None}
    per_label = ""
    if date:
        per_label = f"dal {min(date).day} {MESI_NOME[min(date).month]} {min(date).year}"

    # tipologie (vocabolario controllato; "-" => Non classificato)
    from collections import Counter
    tip = Counter()
    for r in data:
        t = str(col(r, "Tipologia evento") or "-").strip()
        tip["Non classificato" if t in ("-", "", "None") else t] += 1

    write_kv("statistiche-interventi", [
        ("Periodo (provvisorio)", per_label or "n.d."),
        ("Interventi registrati", n_tot),
        ("Ore di intervento totali", ore),
        ("Chilometri percorsi totali", int(km)),
    ], periodo)

    write_rows("interventi-per-tipologia",
               ["tipologia", "numero_interventi"],
               sorted(([k, v] for k, v in tip.items()), key=lambda x: -x[1]),
               periodo)

    # ---------- MEMBRI (solo aggregato, niente nomi) ----------
    ws2 = wb["Interventi membri"]
    d2 = [r for r in ws2.iter_rows(values_only=True)][1:]
    d2 = [r for r in d2 if r and r[0]]
    vol_elenco = len(d2)
    vol_attivi = sum(1 for r in d2 if fnum(r[1]) > 0)
    rap_tot = int(sum(fnum(r[1]) for r in d2))
    rap_comp = int(sum(fnum(r[2]) for r in d2))
    write_kv("statistiche-volontari", [
        ("Volontari in organico", vol_elenco),
        ("Volontari con almeno un intervento", vol_attivi),
        ("Rapporti di intervento totali", rap_tot),
        ("Rapporti compilati", rap_comp),
    ], periodo)

    print("\nRiepilogo (anonimo):")
    print(f"  interventi: {n_tot}, ore {ore}, km {int(km)}, {per_label}")
    print(f"  volontari: {vol_elenco} in organico, {vol_attivi} attivi, {rap_tot} rapporti")
    print("\nNessun dato personale scritto. Il file .xlsx NON va committato.")


if __name__ == "__main__":
    main()
